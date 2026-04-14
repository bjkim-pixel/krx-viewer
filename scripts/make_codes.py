"""
make_codes.py — 로컬에서 1회 실행
────────────────────────────────
KRX 엑셀 파일에서 종목코드 목록을 추출해서
data/codes.json 파일로 저장합니다.
이 파일을 git add → commit → push 하면
GitHub Actions가 읽어서 수급 데이터를 수집합니다.

사용법:
  python scripts/make_codes.py 엑셀파일명.xlsx
"""
import sys, json
from pathlib import Path

def main():
    if len(sys.argv) < 2:
        print("사용법: python scripts/make_codes.py 파일명.xlsx")
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("openpyxl 설치 필요: pip install openpyxl")
        sys.exit(1)

    xl_path = Path(sys.argv[1])
    if not xl_path.exists():
        print(f"파일 없음: {xl_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xl_path)
    ws = wb.active
    headers = [str(c.value or '').strip() for c in ws[1]]

    # 종목코드 컬럼 찾기
    code_col = next((i for i,h in enumerate(headers) if '종목코드' in h or 'ISU_SRT_CD' in h), None)
    if code_col is None:
        print(f"종목코드 컬럼 없음. 헤더: {headers}")
        sys.exit(1)

    # 파일명에서 날짜 추출
    import re
    m = re.search(r'(\d{8})', xl_path.name)
    date = m.group(1) if m else ''

    codes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[code_col] or '').strip().replace("'", '').zfill(6)
        if code and code != '000000':
            codes.append(code)

    out = {
        'codes': codes,
        'date':  date,
        'total': len(codes),
        'source': xl_path.name,
    }
    out_path = Path(__file__).parent.parent / 'data' / 'codes.json'
    out_path.parent.mkdir(exist_ok=True)
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"✅ {len(codes)}개 종목코드 → data/codes.json")
    print(f"   날짜: {date}")
    print()
    print("다음 단계:")
    print("  git add data/codes.json")
    print("  git commit -m '종목 목록 업데이트'")
    print("  git push")
    print("  → GitHub Actions 수동 실행")

if __name__ == '__main__':
    main()
